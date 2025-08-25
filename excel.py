import pandas as pd
import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def generate_comprehensive_facture_excel(data: list[dict], client_full_name: str):
    report_period = datetime.date.today().year

    safe_client_name = "".join(c for c in client_full_name if c.isalnum() or c in (' ', '-', '_')).replace(' ',
                                                                                                           '_').rstrip(
        '_')
    file_name = f"Rapport_Factures_{safe_client_name}_{report_period}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = f"Factures {client_full_name} {report_period}"

    # Styles
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=14)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Définition des couleurs de remplissage
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Vert clair
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Rouge clair

    current_row = 1

    # Informations du client (en-tête)
    if data:
        client_info = data[0]
        client_display_name = f"{client_info['client_nom']} {client_info['client_prenom']}"
        if client_info['client_categorie'] != 'Particulier':
            client_display_name = f"{client_info['client_nom']} (Responsable: {client_info['client_prenom'] if client_info['client_prenom'] else 'N/A'})"

        ws.cell(row=current_row, column=1, value="Client :").font = bold_font
        ws.cell(row=current_row, column=2, value=client_display_name)
        current_row += 1

        # Ajout du numéro de contrat
        ws.cell(row=current_row, column=1, value="N° Contrat :").font = bold_font
        ws.cell(row=current_row, column=2, value=client_info.get('Référence Contrat', 'N/A'))
        current_row += 1

        ws.cell(row=current_row, column=1, value="Adresse :").font = bold_font
        ws.cell(row=current_row, column=2, value=client_info['client_adresse'])
        current_row += 1

        ws.cell(row=current_row, column=1, value="Téléphone :").font = bold_font
        ws.cell(row=current_row, column=2, value=client_info['client_telephone'])
        current_row += 1

        ws.cell(row=current_row, column=1, value="Catégorie Client :").font = bold_font
        ws.cell(row=current_row, column=2, value=client_info['client_categorie'])
        current_row += 1

        ws.cell(row=current_row, column=1, value="Axe Client :").font = bold_font
        ws.cell(row=current_row, column=2, value=client_info['client_axe'])
        current_row += 1

    current_row += 1

    # Ligne de titre du rapport
    ws.cell(row=current_row, column=1,
            value=f"Rapport de Facturation pour la période : {report_period}").font = header_font
    table_headers = [
        'Numéro Facture', 'Date de Planification', 'Date de Facturation', 'Type de Traitement',
        'Etat du Planning', 'Mode de Paiement', 'Détails Paiement', 'Etat de Paiement', 'Montant Facturé'
    ]
    max_cols_for_merge = len(table_headers)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_cols_for_merge)
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
    current_row += 2

    # Tableau des détails de la facture
    # Écrire les en-têtes du tableau
    for col_idx, header in enumerate(table_headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = bold_font
        cell.border = thin_border
    current_row += 1

    if not data:
        ws.cell(row=current_row, column=1,
                value=f"Aucune facture trouvée pour le client '{client_full_name}' pour la période sélectionnée.").border = thin_border
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(table_headers))
        current_row += 1
    else:
        df_invoice_data = pd.DataFrame(data)

        for r_idx, row_dict in enumerate(df_invoice_data.to_dict('records'), start=current_row):
            # Gérer le numéro de facture: afficher "Aucun" si vide ou None
            invoice_number = row_dict.get('Numéro Facture')
            display_invoice_number = invoice_number if invoice_number else "Aucun"

            row_data = [
                display_invoice_number, # Utilisation de la valeur traitée
                row_dict.get('Date de Planification', 'N/A'),
                row_dict.get('Date de Facturation', 'N/A'),
                row_dict.get('Type de Traitement', 'N/A'),
                row_dict.get('Etat du Planning', 'N/A'),
                row_dict.get('Mode de Paiement', 'N/A'),
                '',  # Placeholder for Détails Paiement
                row_dict.get('Etat de Paiement', 'N/A'),
                row_dict.get('Montant Facturé', 'N/A')
            ]

            # Gérer les détails de paiement
            mode_paiement = row_dict.get('Mode de Paiement')
            details_paiement = "N/A"
            date_cheque_obj = row_dict.get('Date de Paiement')
            date_cheque_str = date_cheque_obj.strftime('%Y-%m-%d') if date_cheque_obj else 'N/A'

            if mode_paiement == 'Chèque':
                numero_cheque_str = row_dict.get('Numéro du Chèque', 'N/A')
                etablissement_payeur_str = row_dict.get('Établissement Payeur', 'N/A')
                details_paiement = f"Chèque: {numero_cheque_str} ({date_cheque_str}, {etablissement_payeur_str})"
            elif mode_paiement == 'Virement':
                details_paiement = f"Virement: ({date_cheque_str})"
            elif mode_paiement == 'Mobile Money':
                details_paiement = f"Mobile Money ({date_cheque_str})"
            elif mode_paiement == 'Espèce':
                details_paiement = f"Espèces: ({date_cheque_str})"
            row_data[6] = details_paiement  # Mettre à jour la colonne 'Détails Paiement'


            payment_status = row_dict.get('Etat de Paiement')
            fill_to_apply = None
            if payment_status == 'Payé':
                fill_to_apply = green_fill
            elif payment_status == 'Non payé':
                fill_to_apply = red_fill

            for c_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                if fill_to_apply:
                    cell.fill = fill_to_apply
            current_row += 1

    current_row += 1

    # Calcul et affichage des totaux
    if data:
        df_calc = pd.DataFrame(data)

        grand_total = df_calc['Montant Facturé'].sum()
        ws.cell(row=current_row, column=1, value="Montant Total Facturé sur la période :").font = bold_font
        ws.cell(row=current_row, column=len(table_headers), value=grand_total).font = bold_font
        current_row += 1

        total_paid = df_calc[df_calc['Etat de Paiement'] == 'Payé']['Montant Facturé'].sum()
        ws.cell(row=current_row, column=1, value="Montant Total Payé sur la période :").font = bold_font
        ws.cell(row=current_row, column=len(table_headers), value=total_paid).font = bold_font
        ws.cell(row=current_row, column=len(table_headers)).fill = green_fill
        current_row += 1

        total_unpaid = df_calc[df_calc['Etat de Paiement'] == 'Non payé']['Montant Facturé'].sum()
        ws.cell(row=current_row, column=1, value="Montant Total Impayé sur la période :").font = bold_font
        ws.cell(row=current_row, column=len(table_headers), value=total_unpaid).font = bold_font
        ws.cell(row=current_row, column=len(table_headers)).fill = red_fill
        current_row += 1

        current_row += 1

        # Total de paiement par mode de paiement
        ws.cell(row=current_row, column=1, value="Total de paiement par mode de paiement :").font = bold_font
        current_row += 1
        payment_mode_counts = df_calc.groupby('Mode de Paiement').size().reset_index(name='Nombre de Paiements')
        for _, row in payment_mode_counts.iterrows():
            ws.cell(row=current_row, column=2, value=f"{row['Mode de Paiement']} :").font = bold_font
            ws.cell(row=current_row, column=3, value=row['Nombre de Paiements']).font = bold_font
            current_row += 1
        current_row += 1

        ws.cell(row=current_row, column=1, value="Synthèse par Type de Traitement :").font = bold_font
        current_row += 1

        if 'Type de Traitement' in df_calc.columns:
            total_by_type = df_calc.groupby('Type de Traitement')['Montant Facturé'].sum().reset_index()
            for _, row in total_by_type.iterrows():
                ws.cell(row=current_row, column=2, value=row['Type de Traitement']).font = bold_font
                ws.cell(row=current_row, column=3, value=row['Montant Facturé']).font = bold_font
                current_row += 1
        else:
            ws.cell(row=current_row, column=1,
                    value="Impossible de synthétiser par type de traitement (colonne manquante).")
            current_row += 1

    for i in range(1, len(table_headers) + 1):
        column_letter = get_column_letter(i)
        length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=i)
            if cell.value is not None:
                if isinstance(cell.value, (datetime.date, datetime.datetime)):
                    cell_length = len(cell.value.strftime('%Y-%m-%d'))
                else:
                    cell_length = len(str(cell.value))
                length = max(length, cell_length)
        ws.column_dimensions[column_letter].width = length + 2

    try:
        output = BytesIO()
        wb.save(output)
        # Save to a file in the current directory
        with open(file_name, 'wb') as f:
            f.write(output.getvalue())

        print(f"Fichier '{file_name}' généré avec succès.")
    except Exception as e:
        print(f"Erreur lors de la génération du fichier Excel de la facture : {e}")


def generer_facture_excel(data: list[dict], client_full_name: str, year: int, month: int):
    month_name_fr = datetime.date(year, month, 1).strftime('%B').capitalize()

    safe_client_name = "".join(c for c in client_full_name if c.isalnum() or c in (' ', '-', '_')).replace(' ',
                                                                                                           '_').rstrip(
        '_')
    file_name = f"{safe_client_name}-{month_name_fr}-{year}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = f"Facture {client_full_name} {month_name_fr}"

    # Styles
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=14)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Définition des couleurs de remplissage
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Vert clair
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Rouge clair

    ligneActuelle = 1

    # Informations du client (en-tête)
    if data:
        infoClient = data[0]
        affichageNomClient = f"{infoClient['client_nom']} {infoClient['client_prenom']}"
        if infoClient['client_categorie'] != 'Particulier':
            affichageNomClient = f"{infoClient['client_nom']} (Responsable: {infoClient['client_prenom'] if infoClient['client_prenom'] else 'N/A'})"

        ws.cell(row=ligneActuelle, column=1, value="Client :").font = bold_font
        ws.cell(row=ligneActuelle, column=2, value=affichageNomClient)
        ligneActuelle += 1

        # Ajout du numéro de contrat
        ws.cell(row=ligneActuelle, column=1, value="N° Contrat :").font = bold_font
        ws.cell(row=ligneActuelle, column=2, value=infoClient.get('Référence Contrat', 'N/A'))
        ligneActuelle += 1

        ws.cell(row=ligneActuelle, column=1, value="Adresse :").font = bold_font
        ws.cell(row=ligneActuelle, column=2, value=infoClient['client_adresse'])
        ligneActuelle += 1

        ws.cell(row=ligneActuelle, column=1, value="Téléphone :").font = bold_font
        ws.cell(row=ligneActuelle, column=2, value=infoClient['client_telephone'])
        ligneActuelle += 1

        ws.cell(row=ligneActuelle, column=1, value="Catégorie Client :").font = bold_font
        ws.cell(row=ligneActuelle, column=2, value=infoClient['client_categorie'])
        ligneActuelle += 1

        ws.cell(row=ligneActuelle, column=1, value="Axe Client :").font = bold_font
        ws.cell(row=ligneActuelle, column=2, value=infoClient['client_axe'])
        ligneActuelle += 1

    ligneActuelle += 1

    # Tableau des traitements
    table_headers = [
        'Numéro Facture', 'Date de Planification', 'Date de traitement', 'Traitement concerné',
        'Etat du Planning', 'Mode de Paiement', 'Détails Paiement', 'Etat de Paiement', 'Montant'
    ]
    num_table_cols = len(table_headers)

    # Ligne "Facture du mois de:"
    ws.cell(row=ligneActuelle, column=1, value=f"Facture du mois de : {month_name_fr} {year}").font = header_font
    ws.merge_cells(start_row=ligneActuelle, start_column=1, end_row=ligneActuelle, end_column=num_table_cols)
    ws.cell(row=ligneActuelle, column=1).alignment = Alignment(horizontal='center')
    ligneActuelle += 2

    # Écrire les en-têtes du tableau
    for col_idx, header in enumerate(table_headers, 1):
        cell = ws.cell(row=ligneActuelle, column=col_idx, value=header)
        cell.font = bold_font
        cell.border = thin_border
    ligneActuelle += 1

    if not data:
        ws.cell(row=ligneActuelle, column=1,
                value=f"Aucune facture trouvée pour le client '{client_full_name}' pour ce mois.").border = thin_border
        ws.merge_cells(start_row=ligneActuelle, start_column=1, end_row=ligneActuelle, end_column=len(table_headers))
        ligneActuelle += 1
    else:
        # Convertir en DataFrame pour un traitement plus facile
        df_invoice_data = pd.DataFrame(data)

        for r_idx, row_dict in enumerate(df_invoice_data.to_dict('records'), start=ligneActuelle):
            # Gérer le numéro de facture: afficher "Aucun" si vide ou None
            invoice_number = row_dict.get('Numéro Facture')
            display_invoice_number = invoice_number if invoice_number else "Aucun"

            # Préparer les données de la ligne selon les en-têtes définis
            row_data = [
                display_invoice_number, # Utilisation de la valeur traitée
                row_dict.get('Date de Planification', 'N/A'),
                row_dict.get('Date de traitement', 'N/A'),
                row_dict.get('Traitement (Type)', 'N/A'),
                row_dict.get('Etat traitement', 'N/A'),
                row_dict.get('Mode de Paiement', 'N/A'),
                '',  # Placeholder for Détails Paiement
                row_dict.get('Etat paiement (Payée ou non)', 'N/A'),
                row_dict.get('montant_facture', 'N/A')
            ]

            # Gérer les détails de paiement
            mode_paiement = row_dict.get('Mode de Paiement')
            details_paiement = "N/A"
            date_cheque_obj = row_dict.get('Date de Paiement')
            date_cheque_str = date_cheque_obj.strftime('%Y-%m-%d') if date_cheque_obj else 'N/A'

            if mode_paiement == 'Chèque':
                numero_cheque_str = row_dict.get('Numéro du Chèque', 'N/A')
                etablissement_payeur_str = row_dict.get('Établissement Payeur', 'N/A')
                details_paiement = f"Chèque: {numero_cheque_str} ({date_cheque_str}, {etablissement_payeur_str})"
            elif mode_paiement == 'Virement':
                details_paiement = f"Virement: ({date_cheque_str})"
            elif mode_paiement == 'Mobile Money':
                details_paiement = f"Mobile Money ({date_cheque_str})"
            elif mode_paiement == 'Espèce':
                details_paiement = f"Espèces: ({date_cheque_str})"
            row_data[6] = details_paiement # Mettre à jour la colonne 'Détails Paiement'

            payment_status = row_dict.get('Etat paiement (Payée ou non)')
            fill_to_apply = None
            if payment_status == 'Payé':
                fill_to_apply = green_fill
            elif payment_status == 'Non payé':
                fill_to_apply = red_fill

            for c_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                if fill_to_apply:
                    cell.fill = fill_to_apply
            ligneActuelle += 1

    ligneActuelle += 1

    # Calcul et affichage des totaux
    if data:
        df_calc = pd.DataFrame(data)

        total_by_type_paid = df_calc[df_calc['Etat paiement (Payée ou non)'] == 'Payé'].groupby('Traitement (Type)')[
            'montant_facture'].sum()

        if not total_by_type_paid.empty:
            ws.cell(row=ligneActuelle, column=1, value="Facture total pour :").font = bold_font
            ligneActuelle += 1
            for service_type, total_amount in total_by_type_paid.items():
                ws.cell(row=ligneActuelle, column=2, value=f"{service_type} (Payé)").font = bold_font
                ws.cell(row=ligneActuelle, column=3, value=total_amount).font = bold_font
                ligneActuelle += 1
        else:
            ws.cell(row=ligneActuelle, column=1,
                    value="Aucun montant payé pour les types de traitement ce mois.").font = bold_font
            ligneActuelle += 1

        ligneActuelle += 1

        # Total de paiement par mode de paiement
        ws.cell(row=ligneActuelle, column=1, value="Total de paiement par mode de paiement :").font = bold_font
        ligneActuelle += 1
        payment_mode_counts = df_calc.groupby('Mode de Paiement').size().reset_index(name='Nombre de Paiements')
        for _, row in payment_mode_counts.iterrows():
            ws.cell(row=ligneActuelle, column=2, value=f"{row['Mode de Paiement']} :").font = bold_font
            ws.cell(row=ligneActuelle, column=3, value=row['Nombre de Paiements']).font = bold_font
            ligneActuelle += 1
        ligneActuelle += 1

        grand_total = df_calc['montant_facture'].sum()
        ws.cell(row=ligneActuelle, column=1, value="Montant total des traitements effectués ce mois :").font = bold_font
        ws.cell(row=ligneActuelle, column=3, value=grand_total).font = bold_font
        ligneActuelle += 1

    max_col_for_width = len(table_headers)

    for i in range(1, max_col_for_width + 1):
        column_letter = get_column_letter(i)
        length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=i)
            if cell.value is not None:
                # Gérer les dates pour le calcul de la largeur
                if isinstance(cell.value, (datetime.date, datetime.datetime)):
                    cell_length = len(cell.value.strftime('%Y-%m-%d'))
                else:
                    cell_length = len(str(cell.value))
                length = max(length, cell_length)
        ws.column_dimensions[column_letter].width = length + 2

    try:
        output = BytesIO()
        wb.save(output)
        # Save to a file in the current directory
        with open(file_name, 'wb') as f:
            f.write(output.getvalue())

        print(f"Fichier '{file_name}' généré avec succès.")
    except Exception as e:
        print(f"Erreur lors de la génération du fichier Excel de la facture : {e}")


def generate_traitements_excel(data: list[dict], year: int, month: int):
    month_name_fr = datetime.date(year, month, 1).strftime('%B').capitalize()
    file_name = f"traitements-{month_name_fr}-{year}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = f"Traitements {month_name_fr} {year}"

    # Styles
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=14)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Définition des couleurs de remplissage
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Rouge clair
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Vert clair

    # Titre du rapport
    ws.cell(row=1, column=1, value=f"Rapport des Traitements du mois de {month_name_fr} {year}").font = header_font
    ws.cell(row=1, column=1).alignment = center_align
    # Ajuster le nombre de colonnes fusionnées au nombre réel de colonnes de données
    num_data_cols = len(data[0]) if data else 7 # Fallback if data is empty
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_data_cols)

    # Nombre total de traitements
    total_traitements = len(data)
    ws.cell(row=3, column=1, value=f"Nombre total de traitements ce mois-ci : {total_traitements}").font = bold_font

    # Ligne vide pour la séparation
    ws.cell(row=4, column=1, value="")

    df = pd.DataFrame(data)

    if df.empty:
        ws.cell(row=5, column=1, value="Aucun traitement trouvé pour ce mois.").border = thin_border
        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=num_data_cols)
    else:
        headers = df.columns.tolist()
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_idx, value=header)
            cell.font = bold_font
            cell.border = thin_border # Appliquer la bordure aux en-têtes

        # Itérer sur les données et appliquer la couleur
        for r_idx, row_dict in enumerate(data, start=6):
            for c_idx, col_name in enumerate(headers, 1): # Itérer sur les noms de colonnes pour maintenir l'ordre
                value = row_dict.get(col_name, 'N/A') # Obtenir la valeur par nom de colonne
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border # Appliquer la bordure aux cellules de données

                # Appliquer la couleur si c'est la colonne 'Etat traitement'
                if col_name == 'Etat traitement':
                    if value == 'Effectué':
                        cell.fill = red_fill
                    elif value == 'À venir':
                        cell.fill = green_fill

    max_col_for_width = len(df.columns) if not df.empty else num_data_cols

    for i in range(1, max_col_for_width + 1):
        column_letter = get_column_letter(i)
        length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=i)
            if cell.value is not None:
                if isinstance(cell.value, (datetime.date, datetime.datetime)):
                    cell_length = len(cell.value.strftime('%Y-%m-%d'))
                else:
                    cell_length = len(str(cell.value))
                length = max(length, cell_length)
        ws.column_dimensions[column_letter].width = length + 2

    try:
        output = BytesIO()
        wb.save(output)
        with open(file_name, 'wb') as f:
            f.write(output.getvalue())

        print(f"Fichier '{file_name}' généré avec succès.")
    except Exception as e:
        print(f"Erreur lors de la génération du fichier Excel des traitements : {e}")
